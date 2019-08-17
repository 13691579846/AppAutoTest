"""
------------------------------------
@Time : 2019/8/17 11:57
@Auth : linux超
@File : parseExecl.py
@IDE  : PyCharm
@Motto: Real warriors,dare to face the bleak warning,dare to face the incisive error!
@QQ   : 28174043@qq.com
@GROUP: 878565760
------------------------------------
"""
from openpyxl import load_workbook

from config.globalconf import EXCEL_PATH


class ParseExcel(object):
    def __init__(self, path):
        self.path = path
        self.wd = load_workbook(self.path)

    def get_max_row(self, sheet_name):
        self.ws = self.wd[sheet_name]
        max_row = self.ws.max_row
        return max_row

    def get_max_column(self, sheet_name):
        self.ws = self.wd[sheet_name]
        max_column = self.ws.max_column
        return max_column

    def get_excel_head(self, sheet_name):
        self.ws = self.wd[sheet_name]
        head = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        return head

    def get_row_value(self, sheet_name, row):
        self.ws = self.wd[sheet_name]
        head = self.get_excel_head(sheet_name)
        value = tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        value_dict = dict(zip(head, value))
        return value_dict

if __name__ == '__main__':
    excel = ParseExcel(EXCEL_PATH)
    print(excel.get_row_value("test_login", 2))
